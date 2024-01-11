import openpyxl

from tkinter import filedialog, messagebox

from ..vars import excel_stuff

from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter

from .novel_chapter_tl import NovelChapterTranslated
from .raw_tl_pair import RawTlPair

def write_novelChapterTl_to_excel(
    novelChapterTl: NovelChapterTranslated, 
    receiver_excel_filename: str = None, 
    use_existing_sheetname: str = None, 
    clear_sheet: bool = False, 
    set_headers: bool = False
):
    receiver_excel_filename = filedialog.askopenfilename() if receiver_excel_filename == None else receiver_excel_filename
    wb = openpyxl.load_workbook(
        filename=receiver_excel_filename
    )
    sheet = None
    if use_existing_sheetname == None:
        wb.create_sheet(
            index=0
        )
        sheet = wb[wb.sheetnames[0]]
    elif use_existing_sheetname not in wb.sheetnames:
        wb.create_sheet(
            title=use_existing_sheetname
        )
        sheet = wb[use_existing_sheetname]
    else:
        sheet = wb[use_existing_sheetname]
    
    if clear_sheet:
        old_index = wb.sheetnames.index(sheet.title)
        old_title = sheet.title
        del wb[old_title]
        wb.create_sheet(
            title=old_title, 
            index=old_index
        )
        sheet = wb[old_title]

    # setting headers
    if set_headers:
        headers_to_set = excel_stuff.headers
        for header in headers_to_set:
            if header != None:
                get_cell_from_sheet_0_indexes(
                    sheet=sheet, 
                    x_index=headers_to_set.index(header), 
                    y_index=0
                ).value = header
    
    offset_x = 0
    offset_y = 1

    # chapter metadata
    chapter_metadata: dict[str, RawTlPair] = {
        "novel_title": RawTlPair(
            raw=None, 
            tl=None
        ), 
        "chapter_subtitle": novelChapterTl.chapter_subtitle_tlPair
    }
    chapter_metadata_keys = list(chapter_metadata.keys())
    for meta_attr in chapter_metadata:
        meta_values = [
            meta_attr, 
            chapter_metadata[meta_attr].raw, 
            chapter_metadata[meta_attr].tl
        ]
        for meta_value in meta_values:
            if meta_value == None:
                continue
            get_cell_from_sheet_0_indexes(
                sheet=sheet, 
                x_index=meta_values.index(meta_value), 
                y_index=chapter_metadata_keys.index(meta_attr) + offset_y
            ).value = meta_value
        
    offset_x += len(excel_stuff.metadata_headers)

    # chapter honbun
    chapter_honbun: list[RawTlPair] = novelChapterTl.chapter_honbun_tlPairs
    for tl_pair in chapter_honbun:
        honbun_values = [None, tl_pair.raw, tl_pair.tl]
        for honbun_value in honbun_values:
            if honbun_value == None:
                continue
            get_cell_from_sheet_0_indexes(
                sheet=sheet, 
                x_index=honbun_values.index(honbun_value) + offset_x, 
                y_index=chapter_honbun.index(tl_pair) + offset_y
            ).value = honbun_value
    
    offset_x += len(excel_stuff.honbun_headers)

    # chapter preface
    if novelChapterTl.chapter_preface_tlPairs != None:
        pass

    offset_x += len(excel_stuff.preface_headers)

    # chapter afterword
    if novelChapterTl.chapter_afterword_tlPairs != None:
        pass

    wb.save(
        filename=receiver_excel_filename
    )
    wb.close()

def get_cell_from_sheet_0_indexes(
    sheet: Worksheet, 
    x_index: int, 
    y_index: int
) -> Cell:
    return sheet[f"{get_column_letter(x_index + 1)}{y_index + 1}"]